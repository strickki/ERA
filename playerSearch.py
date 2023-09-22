import requests
import json
import pandas as pd
import tkinter as tk
from openpyxl import load_workbook
from tkinter import filedialog

auth_token = "eyJhbGciOiJSUzUxMiJ9.eyJpc3MiOiJodHRwczovL2R1cHIuZ2ciLCJpYXQiOjE2OTUyMTk2MDEsImp0aSI6IjY5NDA2OTAzNDQiLCJleHAiOjE2OTc4MTE2MDEsInN1YiI6ImMzUnlhV05yT1RjME0wQjVZV2h2Ynk1amIyMD0ifQ.MUCxJ06uYNC-u7zVrTBhmecFxsIEriQyZ8LEAPL9ZB8mYr1eH21ufJAUwvwwWpHu0Ag7NrtKug9l57BtwePIDWdXpq7xsJkfar_BBIwu1QaKsJlxakCjk2ZHdLHRopXMLOE5k_Q9dnmwSys_Qq203tVfyETQEBfWszCdY8zO4E44_N89cg4hUszVh4khxMaXTW41nRg6UyaDbT2DV3nSzQObSiCNsotiRJxQY6zZtMupmHnbzpVMaFycNwFh865Oy0Sy8mjvXi_-hIS0Fdgml6zJLqg28FENlQwsKfaLlT2OzEl70LrG7IFgclUVrolHoDatDv39tgXv7vg5N-yg2A"

# Function to open a file in the system
tk.Tk().withdraw()
def selectExcelFile():
   filepath = filedialog.askopenfilename()
   return filepath

def extractExcelNames(filePath):
    # TODO: Excel functionality 
    workbook = load_workbook(filePath) #filename=filePath)
    sheet = workbook.active
    sheet = workbook['Sheet1']
    firstNames = []
    for i in range(2, sheet.max_row+1):
        firstNames.append(sheet.cell(row=i,column=1).value)
    lastNames = []
    for i in range(2, sheet.max_row+1):
        lastNames.append(sheet.cell(row=i,column=2).value)
    fullNames = []
    for i in range(len(firstNames)):
        fullNames.append(firstNames[i] + " " + lastNames[i])
    
    for i in range(len(fullNames)):
        getData(fullNames[i])

def getData(name):
    headers = {
        "Authorization": "Bearer " + auth_token,
        "Content-type": "application/json"
    }
    data = {
        "limit":10,
        "offset":0,
        "query":name,
        "exclude":[
            
        ],
        "includeUnclaimedPlayers":True,
        "filter":{
            "lat":30.0183494,
            "lng":-95.623966,
            "rating":{
                "maxRating":None,
                "minRating":None
            }
        }
    }
    version = "v1.0"
    url = "https://api.dupr.gg/player/" + version + "/search"
    response = requests.post(url, headers=headers, json=data)
    
    if(response.status_code == 200):
        response = response.json()
    else:
        print("Error with accessing DUPR endpoint")
        exit()
    
    fullName = response['result']['hits'][0]['fullName']
    # lastName = response['result']['hits'][0]['lastName']
    doublesRating = response['result']['hits'][0]['ratings']['doubles']
    singlesRating = response['result']['hits'][0]['ratings']['singles']
    
    print(fullName +  '\'s ' + 'doubles rating is: ' + doublesRating); 
    print(fullName + '\'s ' + 'singles rating is: ' + singlesRating); 
    
# get_data(selectExcelFile())
extractExcelNames("/Users/strickki/Downloads/ERA-Usernames (1).xlsx")
# get_data("rob rhett")